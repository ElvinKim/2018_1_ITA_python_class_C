from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()

ws = wb.active
ws.title = "new sheet name"
ws1 = wb.create_sheet("sheet 1")
ws2 = wb.create_sheet("sheet 2")

ws["A1"] = 10
ws.cell(row=2, column=1, value=20)

for r in range(1, 11):
    for c in range(1, 11):
        ws.cell(row=r, column=c, value=(r-1) * 10 + c)


print("{:-^20}".format(" 여러 셀에 접근 "))
for col_c in ws["C"]:
    print(col_c.value)

for col in ws["A:C"]:
    for cell in col:
        print(cell.value)
    print("-" * 20)


for cell in ws[2]:
    print(cell.value)


for row in ws[1:4]:
    for cell in row:
        print(cell.value)
    print("-" * 20)


for row in ws.iter_rows():
    print(row)

for col in ws.iter_cols(min_row=2, max_row=8, min_col=2, max_col=5):
    print(col[1].value)


print("{:-^20}".format(" 저장 및 로드 "))

loaded_wb = load_workbook("sample.xlsx")
print(loaded_wb.get_sheet_names())
loaded_ws = loaded_wb.get_sheet_by_name("new sheet name")
loaded_ws["A1"] = 200000
loaded_wb.save("sample.xlsx")







