from openpyxl import load_workbook


word_wb = load_workbook("words.xlsx")
word_ws = word_wb.get_sheet_by_name("words")

print("{:-^20}".format(" problem 1 "))

problem1_ws = word_wb.create_sheet("problem1")

for row_i, word in enumerate(word_ws["A"]):
    problem1_ws.cell(row=row_i + 1, column=1, value=word.value)

#word_wb.save("problem1.xlsx")

print("{:-^20}".format(" problem 2 "))

problem2_ws = word_wb.create_sheet("problem2")
for row_i, row in enumerate(word_ws.iter_rows()):
    en_word = row[0].value
    insert_en_word = en_word[0] + "_" * (len(en_word) - 1)
    problem2_ws.cell(row=row_i + 1, column=1, value=row[1].value)
    problem2_ws.cell(row=row_i + 1, column=2, value=insert_en_word)

word_wb.save("problem2.xlsx")

print("{:-^20}".format(" problem 3 "))

problem3_ws = word_wb.create_sheet("problem3")

for row_i, row in enumerate(word_ws.iter_rows()):
    problem3_ws.cell(row=row_i + 1, column=1, value=row[4].value)
    en_sentence = row[3].value

    word_lst = []
    for word in en_sentence.split():
        if word[-1] == "." or word[-1] == ",":
            word_lst.append(word[0] + "_" * (len(word) - 2) + word[-1])
        else:
            word_lst.append(word[0] + "_" * (len(word) - 1))
    problem3_ws.cell(row=row_i + 1, column=2, value=" ".join(word_lst))

word_wb.save("problem3.xlsx")

